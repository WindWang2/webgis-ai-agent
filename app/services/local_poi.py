"""高德全国 POI 入库（GCJ-02 → WGS84）与本地查询服务。

数据流：
- ``python manage.py gd-poi-ingest`` 扫描 ``<LOCAL_GEODATA_DIR>/POI/gd_*_poi.zip``：
  解压 xlsx 到同盘临时文件 → python-calamine ``iter_rows`` 分块流式读取 →
  ``location`` 列 "lng,lat"（GCJ-02）经 ``gcj02_to_wgs84_array``（迭代逆变换，
  精度 ~1m，与 wandergis/coordTransform_py 等主流实现同法）转 WGS84 →
  追加写入 ``<LOCAL_GEODATA_DIR>/gd_pois/gd_pois.gpkg``（单层 pois，GPKG
  自带 R-tree 空间索引 + adcode 属性索引）。
- 副产物：type 含「乡镇级地名」的行回填 yearbook.sqlite 的
  township_centers（乡镇中心点 WGS84），供年鉴乡镇行的空间定位。
- meta.json 边车记录省份/行数，catalog 零扫描。

查询面（工具/路由共用）：query_gd_poi（bbox + 名称/大类/adcode 过滤，
where 下推 SQLite 层，max_features 界定内存），坐标一律 WGS84。
"""
from __future__ import annotations

import itertools
import json
import logging
import re
import shutil
import sqlite3
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

from app.core.config import settings
from app.utils.coord_transform import gcj02_to_wgs84_array

logger = logging.getLogger(__name__)


def _bound_sql(*parts: str) -> str:
    """Join static SQL fragments. Values stay out of the string (bound via ``?``)."""
    return "".join(parts)


_CHUNK_ROWS = 200_000
_RESULT_COLUMNS = [
    "poi_id", "name", "category", "subtype", "typecode",
    "adcode", "adname", "cityname", "pname", "address", "tel",
]


# ── 路径 ──────────────────────────────────────────────────────────────────


def gd_poi_root() -> Path:
    root = (settings.LOCAL_GEODATA_DIR or "").strip()
    return (Path(root).expanduser() if root else Path("data")) / "gd_pois"


def gd_poi_gpkg_path() -> Path:
    return gd_poi_root() / "gd_pois.gpkg"


def gd_poi_source_dir() -> Path:
    root = (settings.LOCAL_GEODATA_DIR or "").strip()
    return (Path(root).expanduser() if root else Path("data")) / "POI"


def gd_poi_available() -> bool:
    return gd_poi_gpkg_path().exists()


def _meta_path() -> Path:
    return gd_poi_root() / "meta.json"


# ── Ingest ────────────────────────────────────────────────────────────────


def _iter_poi_zip_paths(source: Optional[Path] = None) -> List[Path]:
    return sorted((source or gd_poi_source_dir()).glob("gd_*_poi.zip"))


def _province_code(zip_path: Path) -> str:
    """gd_510000_poi.zip → '510000'；多省合并包 gd_61_650000_poi.zip → '61_650000'。"""
    return zip_path.stem.split("_", 1)[1].rsplit("_poi", 1)[0]


_LNG_RE = r"^\s*(-?\d+(?:\.\d+)?)\s*,"
_LAT_RE = r",\s*(-?\d+(?:\.\d+)?)\s*$"


def _parse_location(series: pd.Series) -> Tuple[np.ndarray, np.ndarray]:
    """location "lng,lat" 字符串列 → (lng, lat) float 数组；坏值 NaN。

    用正则抽取而非 split+strip：sheet 尾部可能混有无坐标的对照表段
    （整块全空时 ``.str[1].str.strip()`` 会因全 NaN 抛 AttributeError），
    正则对任意脏值都安全退化为 NaN。
    """
    text = series.astype("string")
    lng = pd.to_numeric(text.str.extract(_LNG_RE, expand=False), errors="coerce")
    lat = pd.to_numeric(text.str.extract(_LAT_RE, expand=False), errors="coerce")
    return lng.to_numpy(dtype=float), lat.to_numpy(dtype=float)


def _to_geo(frame: pd.DataFrame) -> Any:
    """原始块 → WGS84 点 GeoDataFrame（查询契约列 + geometry）。"""
    import geopandas as gpd
    from shapely import points

    lng, lat = _parse_location(frame["location"])
    ok = (
        np.isfinite(lng) & np.isfinite(lat)
        & (np.abs(lng) < 180) & (np.abs(lat) < 90)
    )
    if not ok.all():
        frame, lng, lat = frame[ok], lng[ok], lat[ok]
    wlng, wlat = gcj02_to_wgs84_array(lng, lat)
    type_parts = (
        frame["type"].astype("string").str.split(";", n=2, expand=True)
        .reindex(columns=range(3), fill_value="")
    )
    subtype = (
        type_parts[1].fillna("").astype(str)
        + np.where(type_parts[2] != "", ";" + type_parts[2].fillna(""), "")
    )
    return gpd.GeoDataFrame(
        {
            "poi_id": frame["id"].astype("string").fillna(""),
            "name": frame["name"].astype("string").fillna(""),
            "category": type_parts[0].fillna("").astype(str),
            "subtype": subtype,
            "typecode": frame["typecode"].astype("string").fillna(""),
            # 数值型 adcode（float 化后带 .0）统一清洗为 6 位文本
            "adcode": (
                frame["adcode"].astype("string").fillna("")
                .str.replace(r"\.0$", "", regex=True).str.zfill(6)
            ),
            "adname": frame["adname"].astype("string").fillna(""),
            "cityname": frame["cityname"].astype("string").fillna(""),
            "pname": frame["pname"].astype("string").fillna(""),
            "address": frame["address"].astype("string").fillna(""),
            "tel": frame["tel"].astype("string").fillna(""),
        },
        geometry=points(wlng, wlat),
        crs="EPSG:4326",
    )


def _township_centers(frame: pd.DataFrame) -> List[Tuple[Any, ...]]:
    """type 含「乡镇级地名」的行 → (name, adcode, lng, lat, province, city, county)。"""
    mask = frame["type"].astype("string").str.contains("乡镇级地名", na=False)
    if not mask.any():
        return []
    sub = frame.loc[mask]
    lng, lat = _parse_location(sub["location"])
    wlng, wlat = gcj02_to_wgs84_array(lng, lat)
    out = []
    for rec, wl, wa in zip(sub.to_dict("records"), wlng, wlat):
        out.append((
            rec.get("name"), rec.get("adcode"), float(wl), float(wa),
            rec.get("pname"), rec.get("cityname"), rec.get("adname"),
        ))
    return out


def _write_centers(rows: List[Tuple[Any, ...]]) -> None:
    if not rows:
        return
    from app.services.local_yearbook import yearbook_db_path

    db = yearbook_db_path()
    if not db.exists():
        return
    conn = sqlite3.connect(db)
    try:
        with conn:
            conn.executemany(
                "INSERT OR REPLACE INTO township_centers VALUES (?,?,?,?,?,?,?)",
                rows,
            )
    finally:
        conn.close()


def _dedupe_header(header: List[str]) -> List[str]:
    """重名列（部分省份导出带重复表头）加序号后缀，避免 pandas 静默丢列。"""
    seen: Dict[str, int] = {}
    out: List[str] = []
    for col in header:
        if col in seen:
            seen[col] += 1
            out.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            out.append(col)
    return out


def _flush_chunk(
    chunk: List[Sequence[Any]],
    header: List[str],
    gpkg: Path,
    *,
    append: bool,
) -> Tuple[int, List[Tuple[Any, ...]]]:
    """块 → GPKG（首块建层并补 adcode 索引）→ (写入行数, 乡镇中心点行)。"""
    import pyogrio

    frame = pd.DataFrame(chunk, columns=header)
    gdf = _to_geo(frame)
    centers = _township_centers(frame)
    pyogrio.write_dataframe(gdf, gpkg, layer="pois", driver="GPKG", append=append)
    if not append:
        conn = sqlite3.connect(gpkg)
        try:
            with conn:
                conn.execute('CREATE INDEX IF NOT EXISTS idx_pois_adcode ON "pois"(adcode)')
                conn.execute('CREATE INDEX IF NOT EXISTS idx_pois_category ON "pois"(category)')
        finally:
            conn.close()
    return len(gdf), centers


def _province_prefixes(pcode: str) -> List[str]:
    """包编码 → 省级行政编码前缀：'21_22_230000' → ['21','22','23']。"""
    return [p.zfill(2) for p in pcode.split("_")]


def _delete_province_rows(gpkg: Path, pcode: str) -> None:
    """按省刷新时先删旧数据（GPKG 行 adcode 前缀范围匹配，走索引）。"""
    conn = sqlite3.connect(gpkg)
    try:
        clauses: List[str] = []
        params: List[str] = []
        for p in _province_prefixes(pcode):
            nxt = p[:-1] + chr(ord(p[-1]) + 1)
            clauses.append("(adcode >= ? AND adcode < ?)")
            params.extend((p, nxt))
        with conn:
            conn.execute(
                _bound_sql('DELETE FROM "pois" WHERE ', " OR ".join(clauses)),
                params,
            )
    finally:
        conn.close()


def _iter_xlsx_chunks(
    tmp: Path,
    chunk_rows: int,
) -> Iterable[Tuple[List[str], Iterable[Sequence[Any]]]]:
    """xlsx 成员 → (header, rows-iterable) 序列；覆盖多 sheet 续表（无表头）。

    大省超 xlsx 单表 2^20 行上限被切多 sheet：首 sheet 带表头，续 sheet
    直接是数据行；sheet 尾部可能有整行空白的填充段（坐标过滤时丢弃）。
    """
    try:
        from python_calamine import CalamineWorkbook

        wb = CalamineWorkbook.from_path(str(tmp))
        base_header: Optional[List[str]] = None
        for sheet in wb.sheet_names:
            ws = wb.get_sheet_by_name(sheet)
            it = iter(ws.iter_rows())
            try:
                first_row = [str(c) for c in next(it)]
            except StopIteration:
                continue
            if "location" in first_row and "id" in first_row:
                header = _dedupe_header(first_row)  # 带表头的数据 sheet
                base_header = base_header or header
                yield header, it
            elif base_header is not None:
                yield base_header, itertools.chain([tuple(first_row)], it)  # 无表头续表
            else:
                logger.warning("[gd-poi] %s!%s 表头异常，跳过", tmp.name, sheet)
    except Exception:
        import openpyxl

        wb_xl = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)
        base_header = None
        for sheet in wb_xl.sheetnames:
            ws_xl = wb_xl[sheet]
            row_iter = ([cell.value for cell in row] for row in ws_xl.iter_rows())
            it = iter(row_iter)
            try:
                first_row = [str(c) if c is not None else "" for c in next(it)]
            except StopIteration:
                continue
            if "location" in first_row and "id" in first_row:
                header = _dedupe_header(first_row)
                base_header = base_header or header
                yield header, it
            elif base_header is not None:
                yield base_header, itertools.chain([tuple(first_row)], it)
            else:
                logger.warning("[gd-poi] %s!%s 表头异常，跳过", tmp.name, sheet)


# 江苏 csv 包无表头且比 xlsx 多一个 parent 列：id,parent,name,type,address,location,...
_CSV_NO_HEADER_COLS = [
    "id", "parent", "name", "type", "address", "location", "typecode",
    "pcode", "pname", "citycode", "cityname", "adcode", "adname", "tel",
]


def _iter_csv_chunks(
    tmp: Path,
    chunk_rows: int,
) -> Iterable[Tuple[List[str], Iterable[Sequence[Any]]]]:
    """csv 成员（如江苏包：全引号、无表头、多 parent 列）→ (header, rows) 序列。"""
    read_kwargs = dict(
        dtype=str, chunksize=chunk_rows, keep_default_na=False,
        encoding="utf-8", encoding_errors="replace",
        on_bad_lines="warn",
    )
    probe = pd.read_csv(tmp, nrows=1, header=None, **{
        k: v for k, v in read_kwargs.items() if k != "chunksize"
    })
    first = [str(v) for v in probe.iloc[0].tolist()]
    if "location" in first:  # 带表头的 csv
        names: Optional[List[str]] = first
        header_arg: Any = 0
    else:  # 无表头（真实江苏包）：固定列序
        names = _CSV_NO_HEADER_COLS
        header_arg = None
    header = _dedupe_header([str(c) for c in (names or [])])
    for chunk in pd.read_csv(tmp, header=header_arg, names=names, **read_kwargs):
        chunk.columns = header
        yield header, chunk.itertuples(index=False, name=None)


def ingest_gd_poi(
    source: Optional[Path] = None,
    *,
    force: bool = False,
    provinces: Optional[Sequence[str]] = None,
    progress_cb: Optional[Any] = None,
) -> Dict[str, Any]:
    """gd_*_poi.zip 系列 → gd_pois.gpkg（WGS84）。按 zip 幂等；force 重建
    （指定 provinces 时为按省刷新：只清该省旧行，其余省数据不动）。"""

    zips = _iter_poi_zip_paths(source)
    if provinces:
        want = {str(p).zfill(6) for p in provinces}
        zips = [z for z in zips if _province_code(z) in want]
    if not zips:
        return {"error": "未找到 gd_*_poi.zip（检查 LOCAL_GEODATA_DIR/POI/）"}

    out_dir = gd_poi_root()
    out_dir.mkdir(parents=True, exist_ok=True)
    gpkg = gd_poi_gpkg_path()

    meta_path = _meta_path()
    done: Dict[str, Any] = {}
    if meta_path.exists():
        try:
            done = json.loads(meta_path.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            done = {}
    if force and not provinces:
        # 全量重建：清空整库；按省刷新（force+provinces）只删该省旧行。
        if gpkg.exists():
            gpkg.unlink()
        done = {}

    tmp_dir = out_dir / ".tmp_extract"
    tmp_dir.mkdir(exist_ok=True)
    stats: Dict[str, Any] = {"provinces": {}, "skipped": []}
    # GPKG 层一旦存在，后续所有块一律 append（跨 zip/跨 sheet 累积）。
    layer_exists = gpkg.exists()
    try:
        for zpath in zips:
            pcode = _province_code(zpath)
            if pcode in done.get("provinces", {}) and not force:
                stats["skipped"].append(pcode)
                continue
            if force and gpkg.exists():
                _delete_province_rows(gpkg, pcode)  # 按省刷新：先清旧行
            rows_written = 0
            centers: List[Tuple[Any, ...]] = []
            with zipfile.ZipFile(zpath) as zf:
                member = next(
                    (
                        n for n in zf.namelist()
                        if n.endswith((".xlsx", ".csv"))
                    ),
                    None,
                )
                if member is None:
                    logger.warning("[gd-poi] %s 无 xlsx/csv，跳过", zpath.name)
                    continue
                tmp = tmp_dir / Path(member).name
                with zf.open(member) as src, open(tmp, "wb") as dst:
                    shutil.copyfileobj(src, dst, 4 * 1024 * 1024)
            try:
                if tmp.suffix == ".csv":
                    chunks = _iter_csv_chunks(tmp, _CHUNK_ROWS)
                else:
                    chunks = _iter_xlsx_chunks(tmp, _CHUNK_ROWS)
                for header, rows in chunks:
                    chunk: List[Sequence[Any]] = []
                    for row in rows:
                        if len(row) != len(header):
                            row = tuple(row) + ("",) * (len(header) - len(row))
                        chunk.append(row)
                        if len(chunk) >= _CHUNK_ROWS:
                            n, centers_part = _flush_chunk(
                                chunk, header, gpkg, append=layer_exists,
                            )
                            rows_written += n
                            centers += centers_part
                            layer_exists = True
                            chunk = []
                            if progress_cb:
                                try:
                                    progress_cb(pcode, rows_written)
                                except Exception:  # noqa: BLE001
                                    pass
                    if chunk:
                        n, centers_part = _flush_chunk(
                            chunk, header, gpkg, append=layer_exists,
                        )
                        rows_written += n
                        centers += centers_part
                        layer_exists = True
            finally:
                tmp.unlink(missing_ok=True)
            _write_centers(centers)
            stats["provinces"][pcode] = rows_written
            done["provinces"] = {**done.get("provinces", {}), pcode: rows_written}
            done["generated_at"] = pd.Timestamp.now().isoformat(timespec="seconds")
            meta_path.write_text(
                json.dumps(done, ensure_ascii=False, indent=1), encoding="utf-8",
            )
            logger.info("[gd-poi] %s: %s 行", pcode, rows_written)
        stats["gpkg"] = str(gpkg)
        stats["total_rows"] = sum(stats["provinces"].values())
        return stats
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── 查询面 ────────────────────────────────────────────────────────────────

# 行政区名 → adcode 解析缓存（district.shp 只读一次，名称过滤在毫秒级）。
_district_codes_cache: Dict[str, List[str]] = {}


def _resolve_district_codes(district: str) -> List[str]:
    """行政区名/编码 → adcode 列表（市名→4位前缀，区县名→6位精确）。

    市级用前缀展开（如 成都市 510100 → '5101'，覆盖市内全部区县）；
    区县级精确匹配（如 锦江区 → '510104'）。多省同名区县全部返回
    （OR 连接），与按名查行政区工具的包含语义一致。
    """
    key = str(district).strip()
    if not key:
        return []
    if key in _district_codes_cache:
        return _district_codes_cache[key]
    
    # 纯数字编码直接标准化
    if key.isdigit():
        if len(key) == 6 and key.endswith("0000"):
            out = [key[:2]]
        elif len(key) == 6 and key.endswith("00"):
            out = [key[:4]]
        elif len(key) in (2, 4, 6):
            out = [key]
        else:
            out = [key]
        _district_codes_cache[key] = out
        return out

    codes: List[str] = []
    try:
        from app.tools.local_admin import _load_level

        city = _load_level("city")
        if city is not None and "ct_name" in city.columns:
            m = city[city["ct_name"].astype(str).str.contains(key, na=False, regex=False)]
            for _, r in m.iterrows():
                code = str(r.get("ct_adcode") or r.get("adcode") or "")
                if len(code) >= 4:
                    codes.append(code[:4])
        dist = _load_level("district")
        if dist is not None and "dt_name" in dist.columns:
            m = dist[dist["dt_name"].astype(str).str.contains(key, na=False, regex=False)]
            for _, r in m.iterrows():
                code = str(r.get("dt_adcode") or r.get("adcode") or "")
                if len(code) >= 6:
                    codes.append(code.zfill(6))
    except Exception as exc:  # noqa: BLE001 - 行政区解析失败不阻塞查询
        logger.warning("[gd-poi] district resolve failed for %r: %s", key, exc)
    out = list(dict.fromkeys(codes))[:60]
    _district_codes_cache[key] = out
    return out


def _coerce_polygon(polygon: Any):
    """GeoJSON Polygon/MultiPolygon/FeatureCollection（dict/str/ref 或坐标数组）→ shapely 几何；非法返回 None。"""
    if polygon is None:
        return None
    try:
        from shapely.geometry import shape, box
        from shapely.ops import unary_union

        if hasattr(polygon, "geom_type"):
            return polygon if not polygon.is_empty else None

        if isinstance(polygon, str):
            polygon = polygon.strip()
            if not polygon:
                return None
            if polygon.startswith("ref:"):
                try:
                    from app.services.ref_cache import get_ref_data
                    data = get_ref_data(polygon)
                    if data is not None:
                        return _coerce_polygon(data)
                except Exception:
                    pass
            if polygon.startswith("{") or polygon.startswith("["):
                try:
                    polygon = json.loads(polygon)
                except Exception:
                    return None
            else:
                return None

        if isinstance(polygon, dict):
            t = polygon.get("type")
            if t == "Feature":
                return _coerce_polygon(polygon.get("geometry"))
            if t == "FeatureCollection":
                feats = polygon.get("features") or []
                geoms = [_coerce_polygon(f.get("geometry")) for f in feats if f.get("geometry")]
                valid_geoms = [g for g in geoms if g is not None]
                if not valid_geoms:
                    return None
                u = unary_union(valid_geoms)
                return u if not u.is_empty else None
            if t == "GeometryCollection":
                geoms = [_coerce_polygon(g) for g in polygon.get("geometries", [])]
                valid_geoms = [g for g in geoms if g is not None]
                if not valid_geoms:
                    return None
                u = unary_union(valid_geoms)
                return u if not u.is_empty else None
            if t in ("Polygon", "MultiPolygon"):
                geom = shape(polygon)
                return geom if not geom.is_empty else None
            return None
        elif isinstance(polygon, (list, tuple)):
            if len(polygon) == 4 and all(isinstance(v, (int, float)) for v in polygon):
                minx, miny, maxx, maxy = polygon
                return box(minx, miny, maxx, maxy)
            # 裸坐标环 [[ [lng,lat], ... ], ...] → Polygon
            polygon_dict = {"type": "Polygon", "coordinates": polygon}
            geom = shape(polygon_dict)
            return geom if not geom.is_empty else None
        else:
            return None
    except Exception:  # noqa: BLE001 - 非法 geojson 一律当 None 处理
        return None


_BBOX_RE = re.compile(
    r"^\s*\[?(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\]?\s*$"
)


def _parse_bbox(bbox: Any) -> Optional[Tuple[float, float, float, float]]:
    if isinstance(bbox, (list, tuple)) and len(bbox) == 4:
        try:
            return tuple(float(v) for v in bbox)  # type: ignore[return-value]
        except (TypeError, ValueError):
            return None
    if isinstance(bbox, str):
        m = _BBOX_RE.match(bbox)
        if m:
            return tuple(float(g) for g in m.groups())  # type: ignore[return-value]
    return None


def _sql_like_pattern(value: str) -> str:
    """Escape LIKE metacharacters; result is bound via ``?``, not interpolated."""
    return str(value).replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _sql_like(value: str) -> str:
    """Quoted LIKE fragment for OGR/pyogrio SQL (no bound parameters there)."""
    return "'" + _sql_like_pattern(value).replace("'", "''") + "'"


# 口语子类词 → 高德 taxonomy 片段。高德 subtype 实为「大类;小类」格式
# （高校 = 「学校;高等院校」），LLM 传「大学」时 LIKE 永远 0 命中，
# 兜底查整个 category 又被 LIMIT 截断到索引头部（adcode 最小区县）——
# 实测「成都高校」返回的全是锦江区培训机构。命中表内口语词直接映射。
_SUBTYPE_ALIASES = {
    "大学": "高等院校",
    "高校": "高等院校",
    "universities": "高等院校",
    "university": "高等院校",
    "college": "高等院校",
    "中学": "中学",
    "初中": "中学",
    "高中": "中学",
    "职校": "职业技术学校",
    "中专": "职业技术学校",
    "技校": "职业技术学校",
    "研究所": "科研机构",
    "科研所": "科研机构",
}

# 命中数超过该规模时放弃哈希采样（全国级 category 查询排序代价过高），
# 退回索引顺序头部截断 + note 明示空间偏差。
_SAMPLE_MAX_ROWS = 2_000_000


def gd_poi_catalog() -> Dict[str, Any]:
    available = gd_poi_available()
    out: Dict[str, Any] = {"available": available}
    if available:
        out["gpkg"] = str(gd_poi_gpkg_path())
        if _meta_path().exists():
            meta = json.loads(_meta_path().read_text(encoding="utf-8"))
            out["provinces"] = meta.get("provinces", {})
            out["total_rows"] = sum(meta.get("provinces", {}).values())
    else:
        out["error"] = "gd_pois.gpkg 未生成"
        out["correction_hint"] = "运行 python manage.py gd-poi-ingest 预处理。"
    out["note"] = (
        "query_local_poi(bbox, name_like/category/adcode, ...)；坐标 WGS84；"
        "category 为高德一级分类（餐饮服务/购物服务/…）；subtype 为「大类;小类」"
        "格式的小类段（高校=高等院校、初等=小学/中学/幼儿园），口语别名自动映射。"
    )
    return out


def _adcode_literal(code: str) -> str:
    """#694：adcode → OGR 字面量。只允许数字（adcode 语义即数字编码），
    非法字符剥除后引号翻倍兜底——与 category/subtype/name_like 的转义
    纪律对齐（此前 adcode 是唯一的 f-string 直插注入面）。"""
    digits = "".join(ch for ch in str(code) if ch.isdigit())
    return digits.replace("'", "''")


def _gpkg_blob_envelope(blob: Any) -> Optional[Tuple[float, float, float, float]]:
    """解析 GPKG geometry blob 头的 envelope（G-1/#865）。

    GPKG 二进制头：magic "GP"(2B) + version(1B) + flags(1B) + srs_id(4B LE)
    + 可选 envelope（flags bit1-3 指示维度组合），随后才是 WKB。GDAL
    写 GPKG 默认带 XY envelope——空间过滤无需解析 WKB/无 SpatiaLite。

    Returns (minx, miny, maxx, maxy)；非 GPKG blob 或无 envelope 返回 None。
    """
    import struct

    if not isinstance(blob, (bytes, bytearray, memoryview)) or len(blob) < 40:
        return None
    b = bytes(blob)
    if b[0:2] != b"GP":
        return None
    env_code = (b[3] >> 1) & 0x07
    if env_code == 0:
        return None
    ndoubles = {1: 4, 2: 6, 3: 6, 4: 8}.get(env_code)
    if ndoubles is None:
        return None
    try:
        vals = struct.unpack_from("<%dd" % ndoubles, b, 8)
    except struct.error:
        return None
    minx, maxx, miny, maxy = vals[0], vals[1], vals[2], vals[3]
    return (minx, miny, maxx, maxy)


def _gpkg_rtree_table(conn: sqlite3.Connection) -> Optional[str]:
    """返回 pois 空间列的 R-tree 虚表名；不存在返回 None（#1058）。

    GDAL 写 GPKG 默认为空间列建 ``rtree_<table>_<geom>``；存在性查
    sqlite_master（一次主表 lookup，微秒级）。
    """
    try:
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='rtree_pois_geom'"
        ).fetchone()
    except sqlite3.Error:
        return None
    return row[0] if row else None


def query_gd_poi(
    bbox: Any = None,
    *,
    name_like: Optional[str] = None,
    category: Optional[str] = None,
    subtype: Any = None,
    adcode: Optional[str] = None,
    district: Optional[str] = None,
    polygon: Any = None,
    limit: int = 200,
) -> Dict[str, Any]:
    """空间 + 属性过滤查询。

    空间过滤（三选一，优先级 polygon > bbox > adcode/district）：
    - polygon：WGS84 GeoJSON Polygon/MultiPolygon——bbox 预过滤 + shapely
      精确包含（任意矢量区域，如用户勾选范围）；
    - bbox：WGS84 [minx,miny,maxx,maxy] 或 "w,s,e,n"（矩形，含边界外溢）；
    - district/adcode：行政区精确归属（POI 自带 adcode 字段 + 索引，
      行政区查询首选——无矩形外溢）。
    属性过滤：name_like / category / subtype。G-8（#872）：subtype 支持
    传列表（OR 语义，如 ["小学", "中学"] 对应「中小学」复合词）。
    """
    import pyogrio

    if not gd_poi_available():
        return {
            "error": "本地 POI 库未生成（gd_pois.gpkg 缺失）",
            "correction_hint": "运行 python manage.py gd-poi-ingest；或改用 search_poi 在线工具。",
        }
    parsed = _parse_bbox(bbox)
    poly_geom = _coerce_polygon(polygon)
    if polygon is not None and poly_geom is None:
        return {
            "error": "polygon 需为 WGS84 GeoJSON Polygon/MultiPolygon",
            "correction_hint": "可直接传 get_local_admin_boundary 返回的要素 geometry。",
        }
    district_codes = _resolve_district_codes(district) if district else []
    if district and not district_codes:
        return {
            "error": f"未找到行政区 '{district}'",
            "correction_hint": "试试完整名称（如 '成都市'、'锦江区'），或改用 adcode。",
        }
    if poly_geom is not None:
        parsed = poly_geom.bounds  # polygon 的 bbox 作为索引预过滤范围
    # 全库 5174 万行，过滤条件必须至少命中一个可缩面的维度。
    if (
        parsed is None and not name_like and not adcode
        and not category and not district_codes and poly_geom is None
    ):
        return {
            "error": "bbox、polygon、district、adcode、name_like、category 至少提供一个（subtype 无法独立过滤）",
            "correction_hint": "行政区查询给 district（如 '成都市'）；任意区域给 polygon；"
            "或 bbox/adcode/category。",
        }
    limit = max(1, min(int(limit), 2000))

    # G-8（#872）：subtype 列表 → 多值 OR（每个值独立走别名映射与 LIKE）。
    subtype_list: List[str] = []
    if isinstance(subtype, (list, tuple)):
        subtype_list = [str(s).strip() for s in subtype if str(s).strip()]
    raw_subtype = subtype_list[0] if subtype_list else subtype
    if isinstance(subtype, str) or (subtype is not None and not subtype_list):
        single = _SUBTYPE_ALIASES.get(str(subtype).strip().lower(), subtype) if subtype else subtype
        subtype_list = [single] if single else []
    elif subtype_list:
        subtype_list = [
            _SUBTYPE_ALIASES.get(s.lower(), s) for s in subtype_list
        ]

    # OGR/pyogrio needs an interpolated WHERE; sqlite3 uses bound params.
    # #1058: 谓词必须可下推索引 —— `LOWER(category)=LOWER(?)` 对列套函数，
    # EXPLAIN QUERY PLAN 显示退化为全索引扫描（1.7M 行 COUNT 实测 11.8s），
    # 而 `category = ?` 走 COVERING INDEX（同 COUNT 0.17s，~69×）。中文类目
    # LOWER 恒等，大小写容错只在零命中时用 casefold 变体重试一次兜底。
    def _build_predicates(casefold: bool):
        ogr_c: List[str] = []
        sql_c: List[str] = []
        sql_p: List[Any] = []
        if name_like:
            ogr_c.append(f"name LIKE ('%' || {_sql_like(name_like)} || '%') ESCAPE '\\'")
            sql_c.append("name LIKE ('%' || ? || '%') ESCAPE '\\'")
            sql_p.append(_sql_like_pattern(name_like))
        if category:
            cat = str(category)
            esc = cat.replace(chr(39), chr(39) * 2)
            if casefold:
                ogr_c.append(f"LOWER(category) = LOWER('{esc}')")
                sql_c.append("LOWER(category) = LOWER(?)")
            else:
                ogr_c.append(f"category = '{esc}'")
                sql_c.append("category = ?")
            sql_p.append(cat)
        sub_ogr = None
        sub_sql = None
        sub_sql_p: List[Any] = []
        if subtype_list:
            # G-8（#872）：多值 OR —— 每个值独立 LIKE，括号包裹防与外层 AND 粘连。
            # SQLite LIKE 对 ASCII 默认大小写不敏感（case_sensitive_like 关），
            # 列侧无需再套 LOWER。
            ogr_parts = [
                f"subtype LIKE ('%'||{_sql_like(str(s))}||'%') ESCAPE '\\'"
                for s in subtype_list
            ]
            sql_parts = [
                "subtype LIKE ('%'||?||'%') ESCAPE '\\'"
                for _ in subtype_list
            ]
            sub_ogr = "(" + " OR ".join(ogr_parts) + ")"
            sub_sql = "(" + " OR ".join(sql_parts) + ")"
            sub_sql_p = [_sql_like_pattern(str(s)) for s in subtype_list]
        if adcode:
            code = _adcode_literal(str(adcode).strip())
            if len(code) == 6 and code.endswith("0000"):
                pfx = code[:2]
                nxt = pfx[:-1] + chr(ord(pfx[-1]) + 1)
                ogr_c.append(f"(adcode >= '{pfx}' AND adcode < '{nxt}')")
                sql_c.append("(adcode >= ? AND adcode < ?)")
                sql_p.extend((pfx, nxt))
            elif len(code) == 6 and code.endswith("00"):
                pfx = code[:4]
                nxt = pfx[:-1] + chr(ord(pfx[-1]) + 1)
                ogr_c.append(f"(adcode >= '{pfx}' AND adcode < '{nxt}')")
                sql_c.append("(adcode >= ? AND adcode < ?)")
                sql_p.extend((pfx, nxt))
            elif len(code) in (2, 4):
                nxt = code[:-1] + chr(ord(code[-1]) + 1)
                ogr_c.append(f"(adcode >= '{code}' AND adcode < '{nxt}')")
                sql_c.append("(adcode >= ? AND adcode < ?)")
                sql_p.extend((code, nxt))
            elif len(code) >= 6:  # 区县级 6 位：精确（zfill 兜底补零）
                padded = code.zfill(6)
                ogr_c.append(f"adcode = '{padded}'")
                sql_c.append("adcode = ?")
                sql_p.append(padded)
            elif code:
                # 2/4 位省市级编码：前缀展开。用范围条件而非 LIKE——SQLite 的
                # LIKE 默认大小写不敏感，无法用 idx_pois_adcode（实测 51M 行
                # 全表扫 ~15s），范围条件走 B-tree 索引。
                nxt = code[:-1] + chr(ord(code[-1]) + 1)
                ogr_c.append(f"(adcode >= '{code}' AND adcode < '{nxt}')")
                sql_c.append("(adcode >= ? AND adcode < ?)")
                sql_p.extend((code, nxt))
        for raw in district_codes:
            code = _adcode_literal(str(raw).strip())
            if not code:
                continue
            if len(code) == 4:  # 市级前缀
                nxt = code[:-1] + chr(ord(code[-1]) + 1)
                ogr_c.append(f"(adcode >= '{code}' AND adcode < '{nxt}')")
                sql_c.append("(adcode >= ? AND adcode < ?)")
                sql_p.extend((code, nxt))
            else:  # 区县级精确
                ogr_c.append(f"adcode = '{code}'")
                sql_c.append("adcode = ?")
                sql_p.append(code)
        ogr_w = " AND ".join(ogr_c + ([sub_ogr] if sub_ogr else [])) or None
        sql_w = " AND ".join(sql_c + ([sub_sql] if sub_sql else [])) or None
        sql_w_p = list(sql_p) + sub_sql_p
        base_sql = " AND ".join(sql_c) or None
        return ogr_w, sql_w, sql_w_p, base_sql, list(sql_p)

    where, sql_where, sql_where_params, base_sql, base_sql_params = _build_predicates(False)

    # polygon 精确过滤：bbox 候选量必须大于 limit（过滤后再截断）。
    read_cap = limit
    if poly_geom is not None:
        read_cap = min(20000, max(limit * 5, 5000))

    # 截断均匀采样：pyogrio 的 max_features 按索引顺序取头部——市级
    # category 查询（如成都 科教文化服务 34k 条 limit 2000）会全量落在
    # adcode 最小的区县（实测 2000 条全是锦江区培训机构）。先 COUNT，
    # 超出 read_cap 且规模可控时按 fid 整数哈希采样，覆盖整个查询范围。
    # G-1（#865）：bbox/polygon 路径此前被整体跳过（采样门条件只覆盖
    # "纯属性过滤"）——带 bbox 的查询（本地优先链路的主路径）仍是索引
    # 头部截断，「成都小学」类任务返回点几乎全部来自入库顺序最前的区县，
    # 密度/聚合结论系统性偏斜。空间过滤不能下推 SQLite（无 SpatiaLite
    # 扩展，BuildMBR 不可用）——改为流式扫描 GPKG blob 头的 envelope
    #（纯 Python 解析，无 WKB 开销）取 bbox 内候选 fid。
    total_matched: Optional[int] = None
    sample_fids: Optional[List[int]] = None
    spatial_count_skipped = False

    try:
        conn = sqlite3.connect(
            f"file:{gd_poi_gpkg_path()}?mode=ro", uri=True, timeout=10
        )
        try:
            # 空间候选扫描上限：防止无 R-tree 的库（属性命中百万行）把
            # blob 扫描本身变成瓶颈；超限放弃采样回退旧行为。
            _SPATIAL_SCAN_CAP = 100_000
            # #1058: bbox 命中超过该量级时跳过精确 join COUNT（全国级大
            # bbox 的 COUNT 实测 ~26s），采样照常（均匀、无地域偏差），
            # 计数披露为 None + note。
            _SPATIAL_EXACT_COUNT_MAX_BBOX_ROWS = 2_000_000
            parts: List[str] = []
            params: List[Any] = []
            if sql_where:
                parts.append(sql_where)
                params.extend(sql_where_params)
            if parsed is None:
                # 纯属性路径：谓词直接下推 SQL（既有行为）。
                if not parts:
                    raise ValueError("no predicate for count/sample")
                predicate = " AND ".join(parts)
                total_matched = int(
                    conn.execute(
                        _bound_sql("SELECT COUNT(*) FROM pois WHERE ", predicate),
                        params,
                    ).fetchone()[0]
                )
                if read_cap < total_matched <= _SAMPLE_MAX_ROWS:
                    sample_fids = [
                        int(r[0])
                        for r in conn.execute(
                            _bound_sql(
                                "SELECT fid FROM pois WHERE ",
                                predicate,
                                " ORDER BY (fid * 2654435761 % 2147483647) LIMIT ?",
                            ),
                            [*params, read_cap],
                        )
                    ]
            else:
                # G-1（#865）+ #1058：空间路径优先走 GPKG R-tree（GDAL 为空间
                # 列默认建 rtree_pois_geom 虚表）。候选枚举受 bbox 约束 ——
                # 旧实现按 fid 序（省份-major 摄取序）全表扫描属性匹配行，
                # 100k 扫描配额先被排在前部的省份耗尽（常见类目全国计数
                # >10 万条必然触发），采样被静默禁用、total_matched=None，
                # capped 回退取索引头部 = 系统性地域偏差（实测宽 bbox
                # 餐饮服务 2000/2000 条全来自 5 个重庆 adcode）。
                minx, miny, maxx, maxy = (float(v) for v in parsed)
                predicate = " AND ".join(parts) if parts else None
                rtree_table = _gpkg_rtree_table(conn)
                if rtree_table:
                    bbox_clause = (
                        "r.minx <= ? AND r.maxx >= ? AND r.miny <= ? AND r.maxy >= ?"
                    )
                    bbox_params = [maxx, minx, maxy, miny]
                    base_join = _bound_sql(
                        f"SELECT p.fid FROM {rtree_table} r ",
                        "CROSS JOIN pois p ON p.fid = r.id ",
                        "WHERE ", bbox_clause,
                    )
                    # CROSS JOIN 固定 rtree 驱动（实测 1.4s vs 默认计划 3.6s）。
                    # fid 哈希序 + LIMIT read_cap+1：一次查询同时拿到均匀采样
                    # 与「是否截断」；未截断时 len 即精确 total_matched。
                    sample_sql = base_join
                    sample_params = list(bbox_params)
                    if predicate:
                        sample_sql += " AND " + predicate
                        sample_params += params
                    rows = conn.execute(
                        _bound_sql(
                            sample_sql,
                            " ORDER BY (p.fid * 2654435761 % 2147483647)",
                            f" LIMIT {int(read_cap) + 1}",
                        ),
                        sample_params,
                    ).fetchall()
                    if len(rows) <= read_cap:
                        # 全量命中（≤ limit）：精确计数即行数，无需采样，
                        # 走常规 bbox 读取路径（保持自然序）。
                        total_matched = len(rows)
                    else:
                        bbox_rows = int(conn.execute(
                            _bound_sql(
                                f"SELECT COUNT(*) FROM {rtree_table} r WHERE ",
                                bbox_clause,
                            ),
                            bbox_params,
                        ).fetchone()[0])
                        if bbox_rows <= _SPATIAL_EXACT_COUNT_MAX_BBOX_ROWS:
                            count_sql = base_join
                            count_params = list(bbox_params)
                            if predicate:
                                count_sql += " AND " + predicate
                                count_params += params
                            total_matched = int(conn.execute(
                                _bound_sql(count_sql.replace("SELECT p.fid", "SELECT COUNT(*)", 1)),
                                count_params,
                            ).fetchone()[0])
                        else:
                            # 全国级大 bbox：精确 join COUNT 不成比例地昂贵
                            #（实测 26s）；采样本身仍均匀，只缺计数。
                            total_matched = None
                            spatial_count_skipped = True
                        sample_fids = [int(r[0]) for r in rows[:read_cap]]
                else:
                    # 无 R-tree 的库：保留 GPKG blob envelope 流式扫描 + cap。
                    predicate_full = predicate or "1=1"
                    candidates: List[int] = []
                    scanned = 0
                    capped = False
                    for row in conn.execute(
                        _bound_sql("SELECT fid, geom FROM pois WHERE ", predicate_full),
                        params,
                    ):
                        scanned += 1
                        if scanned > _SPATIAL_SCAN_CAP:
                            capped = True
                            break
                        env = _gpkg_blob_envelope(row[1])
                        if env is None:
                            # 无 envelope（非常规写入）——保守保留为候选。
                            candidates.append(int(row[0]))
                            continue
                        if not (
                            env[0] > maxx or env[2] < minx
                            or env[1] > maxy or env[3] < miny
                        ):
                            candidates.append(int(row[0]))
                    if not capped:
                        total_matched = len(candidates)
                        if read_cap < total_matched:
                            candidates.sort(
                                key=lambda f: (f * 2654435761) % 2147483647
                            )
                            sample_fids = candidates[:read_cap]
            # #1058: 可下推谓词零命中且带 category 时，用 casefold 变体重试
            # 一次（LOWER 列套函数不可下推索引，只作为兜底而非主路径）。
            if (
                category
                and total_matched == 0
                and not sample_fids
                and parsed is None
            ):
                (
                    where_cf, sql_where_cf, sql_where_cf_params,
                    _base_cf, _base_cf_params,
                ) = _build_predicates(True)
                predicate_cf = sql_where_cf
                if predicate_cf:
                    total_matched = int(
                        conn.execute(
                            _bound_sql("SELECT COUNT(*) FROM pois WHERE ", predicate_cf),
                            sql_where_cf_params,
                        ).fetchone()[0]
                    )
                    if read_cap < total_matched <= _SAMPLE_MAX_ROWS:
                        sample_fids = [
                            int(r[0])
                            for r in conn.execute(
                                _bound_sql(
                                    "SELECT fid FROM pois WHERE ",
                                    predicate_cf,
                                    " ORDER BY (fid * 2654435761 % 2147483647) LIMIT ?",
                                ),
                                [*sql_where_cf_params, read_cap],
                            )
                        ]
                    if total_matched:
                        where = where_cf
                        sql_where = sql_where_cf
        finally:
            conn.close()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[gd-poi] count/sample 跳过（%s）", exc)
        total_matched = None
        sample_fids = None

    read_kwargs: Dict[str, Any] = {
        "layer": "pois",
        "columns": _RESULT_COLUMNS,
        "max_features": read_cap,
    }
    if where is not None:
        read_kwargs["where"] = where
    try:
        if sample_fids:
            kwargs = dict(read_kwargs)
            kwargs.pop("max_features", None)
            kwargs["where"] = None
            kwargs["fids"] = sample_fids
            gdf = pyogrio.read_dataframe(str(gd_poi_gpkg_path()), **kwargs)
        elif parsed is not None:
            minx, miny, maxx, maxy = parsed
            if not (-180 <= minx < maxx <= 180 and -90 <= miny < maxy <= 90):
                return {"error": "bbox 数值不合法（需 -180≤minx<maxx≤180、-90≤miny<maxy≤90）"}
            gdf = pyogrio.read_dataframe(str(gd_poi_gpkg_path()), bbox=parsed, **read_kwargs)
        else:
            gdf = pyogrio.read_dataframe(str(gd_poi_gpkg_path()), **read_kwargs)
    except Exception as exc:  # noqa: BLE001
        logger.error("[gd-poi] read failed (where=%r): %s", where, exc)
        return {"error": f"读取 POI 库失败: {exc}"}

    if poly_geom is not None and len(gdf) > 0:
        from shapely import contains_xy, prepare

        xs = gdf.geometry.x.to_numpy()
        ys = gdf.geometry.y.to_numpy()
        prepare(poly_geom)
        keep = contains_xy(poly_geom, xs, ys)
        gdf = gdf[keep].head(limit)

    payload = json.loads(gdf.to_json()) if len(gdf) else {"features": []}
    features = payload.get("features", [])
    out: Dict[str, Any] = {
        "type": "FeatureCollection",
        "features": features,
        "count": len(features),
        "crs_note": "WGS84（已从高德 GCJ-02 转换，迭代逆变换精度 ~1m）",
    }
    if parsed is not None:
        out["bbox"] = list(parsed)
    if total_matched is not None:
        out["total_matched"] = total_matched
    notes: List[str] = []
    _mapped_display = "|".join(subtype_list) if subtype_list else ""
    if raw_subtype and _mapped_display and str(raw_subtype).strip() != _mapped_display:
        notes.append(f"subtype '{raw_subtype}' 已按别名映射为 '{_mapped_display}'")
    if len(features) >= limit:
        out["truncated"] = True
        if sample_fids and total_matched is not None:
            notes.append(
                f"命中 {total_matched} 条超出 limit={limit}，已按 fid 均匀采样返回（空间分布覆盖整个查询范围，而非索引头部单一区县）"
            )
        elif sample_fids and spatial_count_skipped:
            notes.append(
                f"空间命中超过计数预算（bbox 过大），已按 fid 均匀采样返回 limit={limit} 条；"
                "精确 total_matched 未计算——缩小 bbox 或增加属性过滤可获得精确计数。"
            )
        else:
            notes.append(f"结果截断至 limit={limit}，可缩小范围或加过滤。")
    # subtype 零命中：返回该范围内真实子类分布，调用方一轮即可自纠
    # （不再需要猜 taxonomy ——「大学」→ 提示实际值是 学校;高等院校）。
    if raw_subtype and not features:
        hint_sql = base_sql if (
            base_sql and any(k in base_sql for k in ("adcode", "category"))
        ) else None
        if hint_sql is not None:
            try:
                conn = sqlite3.connect(
                    f"file:{gd_poi_gpkg_path()}?mode=ro", uri=True, timeout=10
                )
                try:
                    rows = conn.execute(
                        _bound_sql(
                            "SELECT subtype, COUNT(*) c FROM pois WHERE ",
                            hint_sql,
                            " GROUP BY subtype ORDER BY c DESC LIMIT 12",
                        ),
                        base_sql_params,
                    ).fetchall()
                finally:
                    conn.close()
                if rows:
                    listing = "、".join(f"{s}({c})" for s, c in rows if s)
                    notes.append(
                        f"subtype '{raw_subtype}' 在此范围内无命中。高德子类为「大类;小类」格式，实际分布：{listing}"
                    )
                    out["correction_hint"] = (
                        "请改用上方实际分布中的子类（如 高校 → 高等院校）。"
                    )
            except Exception as exc:  # noqa: BLE001
                logger.warning("[gd-poi] subtype hint 跳过（%s）", exc)
    if notes:
        out["note"] = "；".join(notes)
    return out
