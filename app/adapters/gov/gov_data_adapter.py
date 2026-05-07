"""政府开放数据适配器"""
import logging
import aiohttp
import csv
import io
from datetime import datetime
from typing import Optional
from app.adapters.base import BaseDataAdapter, DataSource
from app.services.explorer.models import (
    RawContent, StructuredData, FieldInfo, SearchContext, DataSourceQualityScore,
)
from app.services.explorer.quality_engine import QualityEngine
from app.core.network import get_base_headers

logger = logging.getLogger(__name__)


class GovDataAdapter(BaseDataAdapter):
    """政府开放数据适配器"""

    name = "gov_data"
    supported_query_types = ["poi_list", "boundary", "statistics"]

    # 已知的政务数据平台
    PLATFORMS = {
        "beijing": {
            "name": "北京市政务数据资源网",
            "search_url": "https://data.beijing.gov.cn/portal/search",
            "base_url": "https://data.beijing.gov.cn",
        },
        "shanghai": {
            "name": "上海市公共数据开放平台",
            "search_url": "https://data.sh.gov.cn/search",
            "base_url": "https://data.sh.gov.cn",
        },
        "guangdong": {
            "name": "广东省政务数据开放平台",
            "search_url": "https://gddata.gd.gov.cn/search",
            "base_url": "https://gddata.gd.gov.cn",
        },
    }

    def __init__(self):
        self.quality_engine = QualityEngine()

    async def discover(self, query: str, context: SearchContext) -> list[DataSource]:
        """探测政府开放数据平台"""
        sources = []

        for platform_id, config in self.PLATFORMS.items():
            try:
                found = await self._search_platform(platform_id, config, query)
                sources.extend(found)
            except Exception as e:
                logger.warning(f"Gov platform {platform_id} search failed: {e}")

        return sources

    async def _search_platform(
        self, platform_id: str, config: dict, query: str
    ) -> list[DataSource]:
        """搜索单个平台"""
        # 简化实现：构造搜索 URL
        search_url = config["search_url"]
        params = {"keyword": query, "page": 1, "size": 10}

        async with aiohttp.ClientSession(headers=get_base_headers()) as session:
            async with session.get(search_url, params=params, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                if resp.status != 200:
                    return []
                data = await resp.json()

        # 解析结果（不同平台格式不同，这里做通用解析）
        items = data.get("data", {}).get("items", []) if isinstance(data, dict) else []
        if not items:
            items = data.get("results", []) if isinstance(data, dict) else []

        sources = []
        for item in items[:5]:  # 最多取 top-5
            source = DataSource(
                id=f"gov_{platform_id}_{item.get('id', 'unknown')}",
                name=item.get("title", "未知数据集"),
                description=item.get("description", ""),
                url=item.get("link", ""),
                format=self._guess_format(item.get("link", "")),
                published_at=self._parse_date(item.get("publish_time", "")),
                estimated_rows=item.get("row_count", 0),
                metadata={"platform": platform_id, "source_url": item.get("link", "")},
            )
            sources.append(source)

        return sources

    async def quick_assess(self, query: str, source: DataSource) -> DataSourceQualityScore:
        """快速质量评估"""
        # 时效性
        temporal = 0.5
        if source.published_at:
            temporal = self.quality_engine.calc_temporal_score(
                "education", source.published_at
            )

        # 主题匹配
        thematic = self.quality_engine.calc_thematic_score(
            user_intent=query,
            dataset_title=source.name,
            dataset_description=source.description,
        )

        # 空间覆盖（政府数据通常是全市范围）
        spatial = 0.3  # 默认覆盖全市，需要过滤

        # 字段完整度（尚未下载，估算）
        field = 0.7 if source.format in ("csv", "xlsx") else 0.4

        # 坐标精度（政府数据通常有地址但无坐标）
        precision = 0.3

        return self.quality_engine.assess_overall(temporal, thematic, spatial, field, precision)

    async def fetch(self, source: DataSource) -> RawContent:
        """下载数据集"""
        if not source.url:
            raise ValueError("Source URL is empty")

        async with aiohttp.ClientSession(headers=get_base_headers()) as session:
            async with session.get(source.url, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                if resp.status != 200:
                    raise RuntimeError(f"Download failed: HTTP {resp.status}")
                data = await resp.read()

        # 大小限制检查
        MAX_SIZE = 50 * 1024 * 1024  # 50MB
        if len(data) > MAX_SIZE:
            raise RuntimeError(f"File too large: {len(data)} bytes > {MAX_SIZE}")

        return RawContent(
            data=data,
            content_type=self._content_type_from_format(source.format),
            encoding=self._detect_encoding(data),
        )

    async def parse(self, raw: RawContent) -> StructuredData:
        """解析 CSV/Excel 为结构化数据"""
        if raw.content_type == "text/csv":
            return self._parse_csv(raw)
        elif raw.content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
            return self._parse_excel(raw)
        else:
            raise ValueError(f"Unsupported format: {raw.content_type}")

    def _parse_csv(self, raw: RawContent) -> StructuredData:
        """解析 CSV"""
        text = raw.data.decode(raw.encoding, errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

        fields = []
        if rows:
            for key in rows[0].keys():
                sample = [r.get(key) for r in rows[:5] if r.get(key)]
                null_count = sum(1 for r in rows if not r.get(key))
                fields.append(FieldInfo(
                    name=key,
                    sample_values=sample,
                    nullable_ratio=round(null_count / len(rows), 4),
                ))

        return StructuredData(rows=rows, fields=fields)

    def _parse_excel(self, raw: RawContent) -> StructuredData:
        """解析 Excel（简化：依赖 openpyxl）"""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl not installed, cannot parse Excel files")

        wb = openpyxl.load_workbook(io.BytesIO(raw.data))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

        fields = []
        if rows:
            for key in headers:
                sample = [r.get(key) for r in rows[:5] if r.get(key)]
                null_count = sum(1 for r in rows if not r.get(key))
                fields.append(FieldInfo(
                    name=key,
                    sample_values=sample,
                    nullable_ratio=round(null_count / len(rows), 4),
                ))

        return StructuredData(rows=rows, fields=fields)

    @staticmethod
    def _guess_format(url: str) -> str:
        if url.endswith(".csv"):
            return "csv"
        elif url.endswith(".xlsx"):
            return "xlsx"
        elif url.endswith(".xls"):
            return "xls"
        elif url.endswith(".json"):
            return "json"
        return "unknown"

    @staticmethod
    def _content_type_from_format(fmt: str) -> str:
        mapping = {
            "csv": "text/csv",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xls": "application/vnd.ms-excel",
            "json": "application/json",
        }
        return mapping.get(fmt, "application/octet-stream")

    @staticmethod
    def _detect_encoding(data: bytes) -> str:
        """检测编码：优先 UTF-8，回退 GBK"""
        try:
            data.decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            return "gbk"

    @staticmethod
    def _parse_date(date_str: str) -> Optional[datetime]:
        """解析日期字符串"""
        if not date_str:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None
